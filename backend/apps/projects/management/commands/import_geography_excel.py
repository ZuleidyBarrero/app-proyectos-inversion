from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.projects.models import Barrio, Comuna, Corregimiento, Vereda


class Command(BaseCommand):
    help = "Importa comunas/barrios y corregimientos/veredas desde un archivo Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Ruta absoluta o relativa al archivo Excel",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"]).expanduser()

        if not excel_path.exists():
            raise CommandError(f"No se encontró el archivo: {excel_path}")

        wb = load_workbook(excel_path)

        self.import_barrios(wb)
        self.import_veredas(wb)

        self.stdout.write(self.style.SUCCESS("Importación finalizada correctamente."))

    def import_barrios(self, wb):
        if "barrios" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("La hoja 'barrios' no existe. Se omite."))
            return

        ws = wb["barrios"]
        headers = [cell.value for cell in ws[1]]

        if headers != ["comuna", "barrio"]:
            raise CommandError(
                "La hoja 'barrios' debe tener exactamente estas columnas: comuna, barrio"
            )

        creados = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            comuna_nombre, barrio_nombre = row

            if not comuna_nombre or not barrio_nombre:
                continue

            comuna_nombre = str(comuna_nombre).strip()
            barrio_nombre = str(barrio_nombre).strip()

            comuna, _ = Comuna.objects.get_or_create(nombre=comuna_nombre)
            _, created = Barrio.objects.get_or_create(
                nombre=barrio_nombre,
                comuna=comuna,
            )
            if created:
                creados += 1

        self.stdout.write(self.style.SUCCESS(f"Barrios importados: {creados}"))

    def import_veredas(self, wb):
        if "veredas" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("La hoja 'veredas' no existe. Se omite."))
            return

        ws = wb["veredas"]
        headers = [cell.value for cell in ws[1]]

        if headers != ["corregimiento", "vereda"]:
            raise CommandError(
                "La hoja 'veredas' debe tener exactamente estas columnas: corregimiento, vereda"
            )

        creados = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            corregimiento_nombre, vereda_nombre = row

            if not corregimiento_nombre or not vereda_nombre:
                continue

            corregimiento_nombre = str(corregimiento_nombre).strip()
            vereda_nombre = str(vereda_nombre).strip()

            corregimiento, _ = Corregimiento.objects.get_or_create(nombre=corregimiento_nombre)
            _, created = Vereda.objects.get_or_create(
                nombre=vereda_nombre,
                corregimiento=corregimiento,
            )
            if created:
                creados += 1

        self.stdout.write(self.style.SUCCESS(f"Veredas importadas: {creados}"))